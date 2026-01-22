import os 
import datetime
import openpyxl
import numpy as np
import h5py
from PIL import Image
import matplotlib.pyplot as plt
import torch
from torchvision import transforms
from torch.utils.data import DataLoader
from torch.nn import functional as F
from tqdm import tqdm

from proxyclip_segmentor import ProxyCLIPSegmentation

import albumentations as A
from dataset.datasets_settings import datasets_settings_test, class_names
from dataset import DWOSMDataset, MultiSenGeDataset, NYCDataset, LoveDADataset, PotsdamDataset
from sam2.get_args import get_args_ov as get_args
import argparse
import sys




KEY_IMG = 'image'
KEY_LBL = 'label'
MASK_CODE = 255

def update_confmat(confmat, pred, gt, n_classes):
    """
    pred, gt: numpy arrays of shape [B,H,W], values in [0..n_classes-1], ignore=0
    confmat: [n_classes, n_classes]
    """
    pred = pred.reshape(-1)
    gt = gt.reshape(-1)
    # ignore 0 (background/ignore) just like your code uses 0
    m = gt != 0
    pred = pred[m]
    gt = gt[m]
    k = n_classes
    cm = np.bincount(k * gt + pred, minlength=k * k).reshape(k, k)
    confmat += cm
    return confmat


COLUMN_INDS = {0:'D1', 1:'E1', 2:'F1', 3:'G1', 4:'H1', 5:'I1', 6:'J1', 7:'K1', 8:'L1', 9:'M1', 10:'N1', 11:'O1', 12:'P1', 13:'Q1', 14:'R1', 15:'S1', 16:'T1',
              17:'U1', 18:'V1', 19:'W1', 20:'X1', 21:'Y1', 22:'Z1'}

def append_experiment_result(file_path, result):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    d = 0
    if sheet['A1'].value is None:
        sheet['A1'] = 'Dataset'
        sheet['B1'] = 'Epoch'
        sheet['C1'] = 'Model'
        for i, name in enumerate(result['names']):
            sheet[COLUMN_INDS[i+d]] = name
        sheet[COLUMN_INDS[i+d+1]] = 'mIoU'
        sheet[COLUMN_INDS[i+d+2]] = 'OA'

    last_row = sheet.max_row+1

    sheet.cell(row=last_row, column=1, value=result['Dataset'])
    sheet.cell(row=last_row, column=2, value=result['Epoch'])
    sheet.cell(row=last_row, column=3, value=result['Model'])
    for i, iou in enumerate(result['IoUs']):
        sheet.cell(row=last_row, column=4+d+i, value=iou)
    sheet.cell(row=last_row, column=5+d+i, value=result['mIoU'])
    sheet.cell(row=last_row, column=6+d+i, value=result['OA'])

    workbook.save(file_path)

def get_class_name_strings(dataset_name):
    cls_dict = class_names[dataset_name]
    name_list = [cls_dict[k][0] for k in cls_dict]
    return name_list

def calculate_acc_from_all(pred, gt, n_classes):
    ious = []
    for i in range(1,n_classes+1):
        pred_ = (pred == i).astype(float)   
        gt_ = (gt == i).astype(float)
        intersection = np.sum(pred_ * gt_)
        union = np.sum(pred_) + np.sum(gt_) - intersection
        iou = intersection / union
        ious.append(iou)
    ious.append(np.mean(ious))
    tmp = pred==gt
    oa = np.sum(tmp[gt>0])/np.sum(gt>0)
    return ious, oa

def main():
    ###### 1 - parameter setting ######
    global args

    # 先解析“分段参数”，并把它们从 sys.argv 里移除，避免 get_args() 报 unknown args
    extra = argparse.ArgumentParser(add_help=False)
    extra.add_argument("--start_idx", type=int, default=0)
    extra.add_argument("--end_idx", type=int, default=-1)   # -1 means run to end
    extra.add_argument("--stats_dir", type=str, default="")
    extra_args, remaining = extra.parse_known_args()

    # 让 get_args() 只看到它认识的参数
    sys.argv = [sys.argv[0]] + remaining
    args = get_args()

    start_idx = extra_args.start_idx
    end_idx = extra_args.end_idx
    stats_dir = extra_args.stats_dir

    assert args.test_only, 'This script is for testing only!'
    if len(args.test_batch_sizes) == 1:
        args.test_batch_sizes = args.test_batch_sizes * len(args.test_dataset_names)
    assert len(args.test_batch_sizes) == len(args.test_dataset_names), 'The number of test batch sizes should be 1 or the same as the number of test datasets!'
    
    for di, args.dataset_name in enumerate(args.test_dataset_names):
        t1 = datetime.datetime.now().replace(microsecond=0)
        patch_size = datasets_settings_test[args.dataset_name]['input_size']
        args.tbs = args.test_batch_sizes[di]
        
        ###### 2 - load data ######
        # datasets
        if args.dataset_name == 'nyc':
            trans_ts = A.Compose([A.CenterCrop(width=patch_size, height=patch_size,)])
        else:
            trans_ts = None 

        # test data
        if args.dataset_name == 'dw':
            test_dataset = DWOSMDataset(datasets_settings_test['dw']['data_path'],
                                        mode=datasets_settings_test['dw']['mode'],
                                        n_bands=datasets_settings_test['dw']['n_channels'],
                                        transform=trans_ts)
        elif args.dataset_name == 'multisenge':
            test_dataset = MultiSenGeDataset(datasets_settings_test['multisenge']['data_path'],
                                            transform=trans_ts)
        elif args.dataset_name == 'osm':
            test_dataset = DWOSMDataset(datasets_settings_test['osm']['data_path'],
                                        mode=datasets_settings_test['osm']['mode'],
                                        n_bands=datasets_settings_test['osm']['n_channels'],
                                        lbl_type='osm')
        elif args.dataset_name == 'nyc':
            test_dataset = NYCDataset(datasets_settings_test['nyc']['data_path'],
                                    n_bands=datasets_settings_test['nyc']['n_channels'],
                                    transform=trans_ts)
        elif args.dataset_name == 'loveda':
            test_dataset = LoveDADataset(datasets_settings_test['loveda']['data_path'],)
        elif args.dataset_name == 'potsdam':
            test_dataset = PotsdamDataset(datasets_settings_test[args.dataset_name]['data_path'],
                                        n_bands=datasets_settings_test[args.dataset_name]['n_channels'],)
        print(f'Test dataset: {args.dataset_name} | Number of samples: {len(test_dataset)} | Patch size: {test_dataset[0][KEY_IMG].shape[-1]}')
        test_loader = DataLoader(test_dataset, batch_size=args.tbs, num_workers=args.num_workers,
                                pin_memory=False, shuffle=False, drop_last=False)
        args.n_batch = len(test_loader)
        args.n_ts = len(test_dataset)

        ###### 3 - set saving path ###### 
        if not os.path.exists(args.save_dir): 
            try:
                os.makedirs(args.save_dir)
            except Exception as inner_e:
                assert os.path.exists(args.save_dir), f'Accuracy directory does not exist and cannot be created: {args.save_dir}. Error: {inner_e}'
        ep = args.ts_wpath.split("epoch=")[-1].split(".")[0]
        pred_fname = f'pred_OV_{args.dataset_name}.xlsx'
        test_save_path = os.path.join(args.save_dir, pred_fname)

        ###### 4 - predict ######
        model = ProxyCLIPSegmentation(clip_type=args.clip_type, 
                                      model_type=args.clip_visual_type, 
                                      vfm_model=args.vfm_model,
                                      name_path=f'ProxyCLIP/configs/cls_{args.dataset_name}.txt',
                                      slide_stride=args.clip_slide_stride, slide_crop=args.clip_slide_crop,
                                      args=args,
                                      )
        
        t0 = datetime.datetime.now().replace(microsecond=0)
        # We'll accumulate confusion matrices instead of storing all predictions.
        conf_clip = None
        conf_weak = None
        conf_fuse = None
        n_classes = None

        # 固定类别数：用数据集的 class_names（最稳）
        n_classes = len(class_names[args.dataset_name])   # loveda 通常是 7
        k = n_classes + 1                                # 加上 0 (ignore)

        conf_clip = np.zeros((k, k), dtype=np.int64)
        conf_weak = np.zeros((k, k), dtype=np.int64)
        conf_fuse = np.zeros((k, k), dtype=np.int64)


        for i, batch in tqdm(enumerate(test_loader), total=args.n_batch):
            if i < start_idx:
                continue
            if end_idx != -1 and i >= end_idx:
                break

            x = batch[KEY_IMG].contiguous().to('cuda', non_blocking=True)
            y = batch[KEY_LBL].contiguous()

            clip_prob = model.predict(x, batch, return_prob=True)
            weak_prob = model.get_weak_model_prediction(batch, return_prob=True)

            weak_prob_max = weak_prob.max(-1)[0].max(-1)[0][:,:,None,None]
            clip_prob_max = clip_prob.max(-1)[0].max(-1)[0][:,:,None,None]
            mask_clip = torch.ones_like(weak_prob_max)*2
            mask_weak = torch.ones_like(weak_prob_max)*2
            select = (weak_prob_max<args.max_thr)*(clip_prob_max>=args.max_thr)
            mask_clip[select] = 3
            mask_weak[select] = 1
            select = (weak_prob_max>=args.max_thr)*(clip_prob_max<args.max_thr)
            mask_clip[select] = 1
            mask_weak[select] = 3
            seg_pred = mask_clip*clip_prob+mask_weak*weak_prob
            seg_pred = seg_pred/seg_pred.sum(dim=1, keepdim=True)

            seg_fuse = seg_pred.argmax(1).cpu().numpy()
            seg_clip = clip_prob.argmax(1).cpu().numpy()
            seg_weak = weak_prob.argmax(1).cpu().numpy()

            y_np = y.numpy()
            # match your original evaluation transform:
            # ys: ignore=0, classes shifted by +1
            ys_eval = np.where(y_np != MASK_CODE, y_np + 1, 0).astype(np.int64)

            # preds are +1 in your original code
            pred_clip_eval = (seg_clip + 1).astype(np.int64)
            pred_weak_eval = (seg_weak + 1).astype(np.int64)
            pred_fuse_eval = (seg_fuse + 1).astype(np.int64)

            # apply ignore mask
            pred_clip_eval[ys_eval == 0] = 0
            pred_weak_eval[ys_eval == 0] = 0
            pred_fuse_eval[ys_eval == 0] = 0


            conf_clip = update_confmat(conf_clip, pred_clip_eval, ys_eval, k)
            conf_weak = update_confmat(conf_weak, pred_weak_eval, ys_eval, k)
            conf_fuse = update_confmat(conf_fuse, pred_fuse_eval, ys_eval, k)

        t1 = datetime.datetime.now().replace(microsecond=0)

        ###### 5 - calculate accuracy from confmat ######
        def confmat_to_iou_oa(conf):
            diag = np.diag(conf).astype(np.float64)
            union = conf.sum(1) + conf.sum(0) - diag
            iou = np.divide(diag, union, out=np.zeros_like(diag), where=union > 0)

            # ignore class 0
            valid = np.arange(conf.shape[0]) != 0
            miou = float(iou[valid].mean()) if valid.any() else 0.0

            # OA over gt>0
            correct = diag[valid].sum()
            total = conf[valid, :].sum()
            oa = float(correct / total) if total > 0 else 0.0
            return iou, miou, oa

        for conf, name in [(conf_clip, 'pc(w.LandSeg)'), (conf_weak, 'LandSeg'), (conf_fuse, 'fusion')]:
            iou, miou, oa = confmat_to_iou_oa(conf)
            print(f'{args.dataset_name} | {name:>15} | mIoU = {miou:.4f} | OA = {oa:.4f}')

        # Save segment stats for later merging (critical for segmented run)
        if stats_dir:
            os.makedirs(stats_dir, exist_ok=True)
            out = os.path.join(
                stats_dir,
                f"{args.dataset_name}_confmat_{start_idx}_{end_idx}.npz"
            )
            np.savez_compressed(out,
                                conf_clip=conf_clip,
                                conf_weak=conf_weak,
                                conf_fuse=conf_fuse)
            print("Saved segment stats:", out)


        t2 = datetime.datetime.now().replace(microsecond=0)
        print(f'Inference time: {t1-t0} | Acc calculating time: {t2-t1}')

if __name__ == '__main__': 
    main()
    


